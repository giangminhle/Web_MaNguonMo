import json
import openpyxl
from math import radians, cos, sin, asin, sqrt
from django.shortcuts import render, redirect, get_object_or_404
from django.contrib.auth import login, logout, authenticate, update_session_auth_hash
from django.contrib.auth.decorators import login_required
from django.contrib.auth.models import User
from django.contrib.auth.forms import PasswordChangeForm
from django.db.models import Q, Prefetch
from decimal import Decimal
from datetime import datetime, timedelta, date
from urllib.parse import urlencode
from django.core.paginator import Paginator 
from django.core.exceptions import PermissionDenied
from django.http import HttpResponse
from django.contrib import messages
from django.contrib.sites.shortcuts import get_current_site
from django.template.loader import render_to_string
from django.utils.http import urlsafe_base64_encode, urlsafe_base64_decode
from django.utils.encoding import force_bytes, force_str
from django.contrib.auth.tokens import default_token_generator
from django.core.mail import EmailMessage

from .models import DiaDiem, SanBong, DatSan, SanPham, ThongBao, ChiTietDichVu, DanhGia, DanhMucSan, TimDoi, HinhAnhDiaDiem, ThongTinGioiThieu
from .forms import DiaDiemForm, SanBongForm, DatSanForm, LoginForm, SignUpForm, SanPhamForm, SuaDonForm, SuaThanhVienForm, CapNhatHoSoForm, GioiThieuForm, ThongBaoForm, TimDoiForm, KhachSuaDonForm

def calculate_distance(lat1, lon1, lat2, lon2):
    R = 6371 
    dlat, dlon = radians(lat2 - lat1), radians(lon2 - lon1)
    a = sin(dlat/2)**2 + cos(radians(lat1)) * cos(radians(lat2)) * sin(dlon/2)**2
    return 2 * R * asin(sqrt(a))

def trang_chu(request):
    query = request.GET.get('q', '')
    cac_dia_diem = DiaDiem.objects.all().order_by('-id')
    if query: cac_dia_diem = cac_dia_diem.filter(Q(ten_dia_diem__icontains=query) | Q(dia_chi__icontains=query)).distinct()
    paginator = Paginator(cac_dia_diem, 4) 
    page_obj = paginator.get_page(request.GET.get('page'))
    ds_thong_bao = ThongBao.objects.all().order_by('-is_pin', '-ngay_dang')[:5]
    return render(request, 'quanlysan/index.html', {'cac_dia_diem': page_obj, 'ds_thong_bao': ds_thong_bao, 'query': query})

def chi_tiet_dia_diem(request, pk):
    dia_diem = get_object_or_404(DiaDiem, pk=pk)
    ds_san_con = dia_diem.ds_san_con.all()
    
    form_admin = None
    if request.user.is_superuser:
        if request.method == 'POST' and 'btn_cap_nhat_admin' in request.POST:
            form_admin = DiaDiemForm(request.POST, request.FILES, instance=dia_diem)
            if form_admin.is_valid():
                dia_diem_sua = form_admin.save()
                for f in request.FILES.getlist('hinh_anh_kem_theo'):
                    HinhAnhDiaDiem.objects.create(dia_diem=dia_diem_sua, hinh_anh=f)
                return redirect('chi_tiet_dia_diem', pk=pk)
        else:
            form_admin = DiaDiemForm(instance=dia_diem)

    user_da_danh_gia = False
    if request.user.is_authenticated and not request.user.is_superuser:
        user_da_danh_gia = DanhGia.objects.filter(dia_diem=dia_diem, khach_hang=request.user).exists()
        if request.method == 'POST' and 'btn_danh_gia' in request.POST and not user_da_danh_gia:
            DanhGia.objects.create(
                dia_diem=dia_diem, khach_hang=request.user, 
                so_sao=int(request.POST.get('so_sao', 5)), binh_luan=request.POST.get('binh_luan', '')
            )
            return redirect('chi_tiet_dia_diem', pk=pk)

    for san in ds_san_con:
        san.lich_hom_nay = DatSan.objects.filter(san=san, ngay_dat=date.today(), trang_thai='DA_DUYET').order_by('gio_bat_dau')
        
    return render(request, 'quanlysan/chi_tiet_dia_diem.html', {
        'dia_diem': dia_diem, 'ds_san_con': ds_san_con, 
        'user_da_danh_gia': user_da_danh_gia, 'form_admin': form_admin,
        'ds_hinh_anh': dia_diem.ds_hinh_anh.all(),
        'ds_danh_gia': dia_diem.ds_danh_gia.all().order_by('-ngay_tao')
    })

def thong_tin_cum_san(request, pk):
    dia_diem = get_object_or_404(DiaDiem, pk=pk)
    ds_san_con = dia_diem.ds_san_con.all()
    so_san_5 = ds_san_con.filter(loai_san=5).count()
    so_san_7 = ds_san_con.filter(loai_san=7).count()
    so_san_11 = ds_san_con.filter(loai_san=11).count()
    prices = [san.gia_tien for san in ds_san_con if san.gia_tien]
    
    gia_min = min(prices) if prices else 0
    gia_max = max(prices) if prices else 0
    gia_min_k = int(gia_min / 1000) if gia_min else 0
    
    context = {
        'dia_diem': dia_diem, 'so_san_5': so_san_5, 'so_san_7': so_san_7, 'so_san_11': so_san_11,
        'gia_min': gia_min, 'gia_max': gia_max, 'gia_min_k': gia_min_k,
        'tong_san': ds_san_con.count(), 'ds_hinh_anh': dia_diem.ds_hinh_anh.all(),
        'ds_danh_gia': dia_diem.ds_danh_gia.all().order_by('-ngay_tao')
    }
    return render(request, 'quanlysan/thong_tin_cum_san.html', context)

@login_required
def xoa_anh_cum_san(request, pk):
    if not request.user.is_superuser: raise PermissionDenied
    anh = get_object_or_404(HinhAnhDiaDiem, pk=pk)
    cum_san_id = anh.dia_diem.id
    anh.delete()
    return redirect('chi_tiet_dia_diem', pk=cum_san_id)

def ban_do_lon(request): return render(request, 'quanlysan/ban_do.html', {'cac_dia_diem': DiaDiem.objects.all(), 'categories': DanhMucSan.objects.all()})

def dang_nhap(request):
    if request.method == 'POST':
        form = LoginForm(request.POST)
        if form.is_valid():
            user = authenticate(request, username=form.cleaned_data.get('username'), password=form.cleaned_data.get('password'))
            if user: login(request, user); return redirect('home')
            form.add_error(None, 'Tên đăng nhập hoặc mật khẩu không chính xác.')
    else: form = LoginForm()
    return render(request, 'quanlysan/login.html', {'form': form})

def dang_ky(request):
    if request.method == 'POST':
        form = SignUpForm(request.POST)
        if form.is_valid():
            user = form.save(commit=False); user.is_active = False; user.save()
            current_site = get_current_site(request)
            mail_subject = 'Kích hoạt tài khoản Sân Bóng 333'
            message = render_to_string('quanlysan/email_kich_hoat.html', {'user': user, 'domain': current_site.domain, 'uid': urlsafe_base64_encode(force_bytes(user.pk)), 'token': default_token_generator.make_token(user)})
            email = EmailMessage(mail_subject, message, to=[form.cleaned_data.get('email')]); email.content_subtype = "html"; email.send()
            return render(request, 'quanlysan/thong_bao_email.html', {'title': 'Đăng ký thành công!', 'message': 'Vui lòng kiểm tra email để kích hoạt tài khoản.'})
    else: form = SignUpForm()
    return render(request, 'quanlysan/signup.html', {'form': form})

def kich_hoat_tai_khoan(request, uidb64, token):
    try: uid = force_str(urlsafe_base64_decode(uidb64)); user = User.objects.get(pk=uid)
    except(TypeError, ValueError, OverflowError, User.DoesNotExist): user = None
    if user is not None and default_token_generator.check_token(user, token):
        user.is_active = True; user.save(); login(request, user)
        return render(request, 'quanlysan/thong_bao_email.html', {'title': 'Thành công!', 'message': 'Tài khoản đã được kích hoạt.', 'is_success': True})
    return render(request, 'quanlysan/thong_bao_email.html', {'title': 'Lỗi!', 'message': 'Link xác nhận không hợp lệ.'})

def dang_xuat(request): logout(request); return redirect('home')

@login_required
def ho_so_ca_nhan(request): return render(request, 'quanlysan/ho_so.html')

@login_required
def doi_mat_khau(request):
    if request.method == 'POST':
        form = PasswordChangeForm(request.user, request.POST)
        if form.is_valid(): user = form.save(); update_session_auth_hash(request, user); return redirect('ho_so')
    else: form = PasswordChangeForm(request.user)
    return render(request, 'quanlysan/doi_mat_khau.html', {'form': form})

@login_required
def lich_su_dat(request): 

    danh_sach_don = DatSan.objects.filter(khach_hang=request.user).order_by('-created_at')
    

    paginator = Paginator(danh_sach_don, 6) 
    page_number = request.GET.get('page')
    ds_don = paginator.get_page(page_number)
    
    return render(request, 'quanlysan/lich_su.html', {'ds_don': ds_don})

@login_required
def khach_huy_don(request, pk):
    don = get_object_or_404(DatSan, pk=pk, khach_hang=request.user, trang_thai='CHO_DUYET')
    don.trang_thai = 'TU_CHOI'
    don.save()
    return redirect('lich_su_dat')

@login_required
def khach_sua_don(request, pk):
    don = get_object_or_404(DatSan, pk=pk, khach_hang=request.user, trang_thai='CHO_DUYET')
    if request.method == 'POST':
        form = KhachSuaDonForm(request.POST, instance=don)
        if form.is_valid():
            form.save()
            return redirect('lich_su_dat')
    else:
        form = KhachSuaDonForm(instance=don)
    return render(request, 'quanlysan/form_chung.html', {'form': form, 'title': 'Sửa Thông Tin Liên Hệ'})

@login_required
def trang_tim_doi(request):
    ds_keo = TimDoi.objects.filter(trang_thai='DANG_TIM', ngay_da__gte=date.today()).order_by('ngay_da', 'gio_bat_dau')
    if request.method == 'POST':
        form = TimDoiForm(request.POST)
        if form.is_valid():
            keo = form.save(commit=False)
            keo.nguoi_dang = request.user
            keo.save()
            return redirect('trang_tim_doi')
    else:
        form = TimDoiForm()
    return render(request, 'quanlysan/tim_doi.html', {'ds_keo': ds_keo, 'form': form})

@login_required
def sua_keo(request, pk):
    keo = get_object_or_404(TimDoi, pk=pk, nguoi_dang=request.user, trang_thai='DANG_TIM')
    if request.method == 'POST':
        form = TimDoiForm(request.POST, instance=keo)
        if form.is_valid():
            form.save()
            return redirect('trang_tim_doi')
    else:
        form = TimDoiForm(instance=keo)
    return render(request, 'quanlysan/form_chung.html', {'form': form, 'title': 'Sửa Kèo Giao Lưu'})

@login_required
def huy_keo(request, pk):
    keo = get_object_or_404(TimDoi, pk=pk, nguoi_dang=request.user, trang_thai='DANG_TIM')
    keo.delete()
    return redirect('trang_tim_doi')

@login_required
def nhan_keo(request, pk):
    keo = get_object_or_404(TimDoi, pk=pk)
    if keo.trang_thai == 'DANG_TIM' and keo.nguoi_dang != request.user:
        q = urlencode({'ngay': keo.ngay_da, 'gio': keo.gio_bat_dau.strftime('%H:%M'), 'phut': keo.thoi_luong, 'keo_id': keo.id})
        return redirect(f"{redirect('dat_san', pk=keo.san_bong.id).url}?{q}")
    return redirect('trang_tim_doi')

@login_required
def dat_san(request, pk):
    san_con = get_object_or_404(SanBong, pk=pk)
    ds_san_pham = SanPham.objects.filter(dia_diem=san_con.dia_diem, so_luong__gt=0)
    error_msg = None
    initial_data = {'ho_ten': f"{request.user.last_name} {request.user.first_name}"}
    if request.GET.get('ngay'): initial_data.update({'ngay_dat': request.GET.get('ngay'), 'gio_bat_dau': request.GET.get('gio'), 'thoi_luong': request.GET.get('phut')})

    if request.method == 'POST':
        form = DatSanForm(request.POST)
        if form.is_valid():
            ngay, gio_bd, phut = form.cleaned_data['ngay_dat'], form.cleaned_data['gio_bat_dau'], form.cleaned_data['thoi_luong']
            gio_kt = (datetime.combine(date.today(), gio_bd) + timedelta(minutes=phut)).time()
            trung = DatSan.objects.filter(san=san_con, ngay_dat=ngay, trang_thai__in=['CHO_DUYET', 'DA_DUYET']).filter(Q(gio_bat_dau__lt=gio_kt) & Q(gio_bat_dau__gte=gio_bd) | Q(gio_bat_dau__lte=gio_bd, thoi_luong__gt=0))
            if any((datetime.combine(ngay, gio_bd) < (datetime.combine(ngay, d.gio_bat_dau) + timedelta(minutes=d.thoi_luong)) and (datetime.combine(ngay, gio_bd) + timedelta(minutes=phut)) > datetime.combine(ngay, d.gio_bat_dau)) for d in trung):
                error_msg = f"Giờ {gio_bd.strftime('%H:%M')} ngày {ngay.strftime('%d/%m/%Y')} đã bận!"
            else:
                danh_mua, tong_dv = [], Decimal(0)
                for sp in ds_san_pham:
                    sl = int(request.POST.get(f'sp_{sp.id}', '0'))
                    if sl > sp.so_luong: error_msg = f"'{sp.ten_sp}' hết hàng!"; break
                    if sl > 0: danh_mua.append((sp, sl)); tong_dv += sp.gia * Decimal(sl)
                if not error_msg:
                    don = form.save(commit=False); don.san, don.khach_hang = san_con, request.user
                    don.tong_tien = (san_con.gia_tien / Decimal(60)) * Decimal(phut); don.save()
                    for sp, sl in danh_mua: ChiTietDichVu.objects.create(don_dat=don, san_pham=sp, so_luong=sl)
                    don.tong_tien += tong_dv; don.tien_coc = don.tong_tien * Decimal('0.3'); don.save()
                    if request.GET.get('keo_id'):
                        k = get_object_or_404(TimDoi, pk=request.GET.get('keo_id'))
                        k.doi_thu = request.user
                        k.trang_thai = 'DA_CAP'
                        k.save()
                    return render(request, 'quanlysan/thanh_cong.html', {'don': don})
    return render(request, 'quanlysan/dat_san.html', {'form': DatSanForm(initial=initial_data), 'san': san_con, 'error_msg': error_msg, 'ds_san_pham': ds_san_pham})

@login_required
def ds_san_pham(request):
    if not request.user.is_staff: raise PermissionDenied
    query = request.GET.get('q', '')
    if request.method == 'POST':
        form = SanPhamForm(request.POST, request.FILES)
        if form.is_valid(): form.save(); messages.success(request, "Thêm sản phẩm thành công!"); return redirect('ds_san_pham')
    else: form = SanPhamForm()
    if query:
        san_pham_qs = SanPham.objects.filter(ten_sp__icontains=query)
        danh_sach_chi_nhanh = DiaDiem.objects.filter(kho_hang__ten_sp__icontains=query).prefetch_related(Prefetch('kho_hang', queryset=san_pham_qs)).distinct().order_by('-id')
    else:
        danh_sach_chi_nhanh = DiaDiem.objects.prefetch_related('kho_hang').all().order_by('-id')
    return render(request, 'quanlysan/ds_san_pham.html', {'cac_dia_diem': Paginator(danh_sach_chi_nhanh, 2).get_page(request.GET.get('page')), 'form': form, 'query': query})

@login_required
def xuat_excel_san_pham(request):
    if not request.user.is_staff: raise PermissionDenied
    response = HttpResponse(content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
    response['Content-Disposition'] = 'attachment; filename="Kho_Hang_San333.xlsx"'
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Kho Hang"
    columns = ['ID SP (K SỬA)', 'Cụm Sân', 'Tên Sản Phẩm', 'Giá Bán', 'Số Lượng Tồn']
    ws.append(columns)
    for cell in ws[1]: cell.font = openpyxl.styles.Font(bold=True)
    for sp in SanPham.objects.select_related('dia_diem').all():
        ws.append([sp.id, sp.dia_diem.ten_dia_diem, sp.ten_sp, float(sp.gia), sp.so_luong])
    wb.save(response)
    return response

@login_required
def nhap_excel_san_pham(request):
    if not request.user.is_staff: raise PermissionDenied
    if request.method == 'POST' and request.FILES.get('excel_file'):
        file = request.FILES['excel_file']
        if not file.name.endswith('.xlsx'):
            messages.error(request, "Chỉ chấp nhận file .xlsx")
            return redirect('ds_san_pham')
        try:
            wb = openpyxl.load_workbook(file)
            ws = wb.active
            preview_list = []
            for row in ws.iter_rows(min_row=2, values_only=True):
                sp_id = row[0]
                ten_san_excel = str(row[1]).strip() if row[1] else ""
                gia_moi = row[3]
                sl_moi = row[4]
                
                if sp_id:
                    try:
                        sp = SanPham.objects.get(id=sp_id)
                        dia_diem_id_moi = sp.dia_diem.id
                        cum_san_hien_thi = sp.dia_diem.ten_dia_diem
                        
                        # KIỂM TRA: Nếu tên sân trong Excel khác với DB -> Cảnh báo chuyển kho
                        if ten_san_excel and ten_san_excel.lower() != sp.dia_diem.ten_dia_diem.lower():
                            san_moi = DiaDiem.objects.filter(ten_dia_diem__iexact=ten_san_excel).first()
                            if san_moi:
                                dia_diem_id_moi = san_moi.id
                                cum_san_hien_thi = f"{san_moi.ten_dia_diem} (Đổi cụm)"
                            else:
                                messages.error(request, f"Lỗi ở ID {sp_id}: Không tìm thấy sân nào có tên '{ten_san_excel}'. Vui lòng nhập đúng chính tả.")
                                return redirect('ds_san_pham')

                        preview_list.append({
                            'id': sp.id,
                            'ten_sp': sp.ten_sp,
                            'cum_san': cum_san_hien_thi,
                            'dia_diem_id_moi': dia_diem_id_moi,
                            'gia_cu': float(sp.gia),
                            'gia_moi': float(gia_moi) if gia_moi is not None else float(sp.gia),
                            'sl_cu': sp.so_luong,
                            'sl_moi': int(sl_moi) if sl_moi is not None else sp.so_luong
                        })
                    except SanPham.DoesNotExist: continue
                    
            if not preview_list:
                messages.error(request, "Không tìm thấy dữ liệu hợp lệ trong file Excel.")
                return redirect('ds_san_pham')
                
            json_data = json.dumps(preview_list)
            return render(request, 'quanlysan/xac_nhan_excel.html', {'preview_list': preview_list, 'json_data': json_data})
        except Exception as e: messages.error(request, f"Lỗi đọc file: {str(e)}")
    return redirect('ds_san_pham')

@login_required
def xac_nhan_excel(request):
    if not request.user.is_staff: raise PermissionDenied
    if request.method == 'POST':
        json_data = request.POST.get('excel_data')
        if json_data:
            try:
                data = json.loads(json_data)
                count = 0
                ds_nhap = []
                ds_xuat = []
                
                for item in data:
                    try:
                        sp = SanPham.objects.get(id=item['id'])
                        sl_cu = sp.so_luong
                        sl_moi = int(item['sl_moi'])
                        sl_thay_doi = sl_moi - sl_cu
                        
                        sp.gia = Decimal(str(item['gia_moi']))
                        sp.so_luong = sl_moi
                        
                        # CẬP NHẬT: Đổi Cụm Sân nếu quản lý có sửa trong file Excel
                        if 'dia_diem_id_moi' in item and item['dia_diem_id_moi'] != sp.dia_diem.id:
                            sp.dia_diem_id = item['dia_diem_id_moi']
                            
                        sp.save()
                        count += 1
                        
                        info = {
                            'ten_sp': sp.ten_sp,
                            'cum_san': sp.dia_diem.ten_dia_diem,
                            'gia': sp.gia,
                            'sl_thay_doi': abs(sl_thay_doi) 
                        }
                        
                        if sl_thay_doi > 0: ds_nhap.append(info)
                        elif sl_thay_doi < 0: ds_xuat.append(info)
                            
                    except SanPham.DoesNotExist: continue
                        
                messages.success(request, f"Đã cập nhật hệ thống thành công {count} mặt hàng!")
                return render(request, 'quanlysan/hoa_don_nhap.html', {'nguoi_nhap': request.user, 'ngay_nhap': datetime.now(), 'ds_nhap': ds_nhap, 'ds_xuat': ds_xuat})
            except Exception as e:
                messages.error(request, f"Lỗi khi lưu dữ liệu: {str(e)}")
    return redirect('ds_san_pham')

@login_required
def sua_san_pham(request, pk):
    if not request.user.is_staff: raise PermissionDenied
    sp = get_object_or_404(SanPham, pk=pk)
    if request.method == 'POST':
        form = SanPhamForm(request.POST, request.FILES, instance=sp)
        if form.is_valid(): form.save(); return redirect('ds_san_pham')
    else: form = SanPhamForm(instance=sp)
    return render(request, 'quanlysan/form_chung.html', {'form': form, 'title': f'Sửa: {sp.ten_sp}'})

@login_required
def xoa_san_pham(request, pk):
    if not request.user.is_staff: raise PermissionDenied
    get_object_or_404(SanPham, pk=pk).delete(); return redirect('ds_san_pham')

@login_required
def quan_ly_don(request):
    if not request.user.is_staff: raise PermissionDenied
    query = request.GET.get('q', '')
    ds_don = DatSan.objects.all().order_by('-created_at')
    if query:
        ds_don = ds_don.filter(Q(ho_ten__icontains=query) | Q(san__ten_san__icontains=query))
    return render(request, 'quanlysan/quan_ly_don.html', {'ds_don': ds_don, 'query': query})

@login_required
def sua_don(request, pk):
    if not request.user.is_staff: raise PermissionDenied
    don = get_object_or_404(DatSan, pk=pk)
    if request.method == 'POST':
        form = SuaDonForm(request.POST, instance=don)
        if form.is_valid(): form.save(); return redirect('quan_ly_don')
    else: form = SuaDonForm(instance=don)
    return render(request, 'quanlysan/form_chung.html', {'form': form, 'title': f'Sửa Đơn: {don.ho_ten}'})

@login_required
def duyet_don(request, pk, trang_thai):
    if not request.user.is_staff: raise PermissionDenied
    d = get_object_or_404(DatSan, pk=pk)
    if trang_thai == 'DA_DUYET' and d.trang_thai == 'CHO_DUYET':
        for ct in d.chi_tiet_dv.all(): ct.san_pham.so_luong = max(0, ct.san_pham.so_luong - ct.so_luong); ct.san_pham.save()
    elif trang_thai == 'TU_CHOI' and d.trang_thai == 'DA_DUYET':
        for ct in d.chi_tiet_dv.all(): ct.san_pham.so_luong += ct.so_luong; ct.san_pham.save()
    d.trang_thai = trang_thai; d.save(); return redirect('quan_ly_don')

@login_required
def quan_ly_thong_bao(request):
    if not request.user.is_superuser: raise PermissionDenied
    if request.method == 'POST':
        ThongBao.objects.create(tieu_de=request.POST.get('tieu_de'), noi_dung=request.POST.get('noi_dung'), is_pin=request.POST.get('is_pin')=='on')
        return redirect('quan_ly_thong_bao')
    return render(request, 'quanlysan/quan_ly_thong_bao.html', {'ds_tb': ThongBao.objects.all().order_by('-ngay_dang')})

@login_required
def sua_thong_bao(request, pk):
    if not request.user.is_superuser: raise PermissionDenied
    tb = get_object_or_404(ThongBao, pk=pk)
    if request.method == 'POST':
        form = ThongBaoForm(request.POST, instance=tb)
        if form.is_valid(): 
            form.save()
            return redirect('quan_ly_thong_bao')
    else: 
        form = ThongBaoForm(instance=tb)
    return render(request, 'quanlysan/form_chung.html', {'form': form, 'title': 'Sửa Thông Báo'})

@login_required
def xoa_thong_bao(request, pk):
    if not request.user.is_superuser: raise PermissionDenied
    get_object_or_404(ThongBao, pk=pk).delete(); return redirect('quan_ly_thong_bao')

@login_required
def them_dia_diem(request):
    if not request.user.is_superuser: raise PermissionDenied
    if request.method == 'POST':
        form = DiaDiemForm(request.POST, request.FILES)
        if form.is_valid(): 
            dia_diem_moi = form.save()
            for f in request.FILES.getlist('hinh_anh_kem_theo'): HinhAnhDiaDiem.objects.create(dia_diem=dia_diem_moi, hinh_anh=f)
            return redirect('home')
    else: form = DiaDiemForm()
    return render(request, 'quanlysan/them_dia_diem.html', {'form': form, 'title': 'Thêm Cụm Sân'})

@login_required
def sua_dia_diem(request, pk):
    if not request.user.is_superuser: raise PermissionDenied
    dd = get_object_or_404(DiaDiem, pk=pk)
    if request.method == 'POST':
        form = DiaDiemForm(request.POST, request.FILES, instance=dd)
        if form.is_valid(): 
            dia_diem_sua = form.save()
            for f in request.FILES.getlist('hinh_anh_kem_theo'): HinhAnhDiaDiem.objects.create(dia_diem=dia_diem_sua, hinh_anh=f)
            return redirect('home')
    else: form = DiaDiemForm(instance=dd)
    return render(request, 'quanlysan/them_dia_diem.html', {'form': form, 'title': f'Sửa: {dd.ten_dia_diem}'})

@login_required
def xoa_dia_diem(request, pk):
    if not request.user.is_superuser: raise PermissionDenied
    get_object_or_404(DiaDiem, pk=pk).delete(); return redirect('home')

@login_required
def them_san_con(request):
    if not request.user.is_superuser: raise PermissionDenied
    if request.method == 'POST':
        form = SanBongForm(request.POST, request.FILES)
        if form.is_valid(): form.save(); return redirect('home')
    else: form = SanBongForm()
    return render(request, 'quanlysan/form_chung.html', {'form': form, 'title': 'Thêm Sân Con'})

@login_required
def sua_san(request, pk):
    if not request.user.is_superuser: raise PermissionDenied
    san = get_object_or_404(SanBong, pk=pk)
    if request.method == 'POST':
        form = SanBongForm(request.POST, request.FILES, instance=san)
        if form.is_valid(): form.save(); return redirect('home')
    else: form = SanBongForm(instance=san)
    return render(request, 'quanlysan/sua_san.html', {'form': form, 'san': san})

@login_required
def quan_ly_thanh_vien(request):
    if not request.user.is_superuser: raise PermissionDenied
    query = request.GET.get('q', '')
    users = User.objects.filter(is_staff=False).order_by('-date_joined')
    if query:
        users = users.filter(Q(username__icontains=query) | Q(first_name__icontains=query) | Q(last_name__icontains=query) | Q(email__icontains=query))
    return render(request, 'quanlysan/quan_ly_thanh_vien.html', {'users': users, 'query': query})

@login_required
def xoa_thanh_vien(request, pk):
    if not request.user.is_superuser: raise PermissionDenied
    u = get_object_or_404(User, pk=pk); u.delete() if not u.is_staff else None; return redirect('quan_ly_thanh_vien')

@login_required
def sua_thanh_vien(request, pk):
    if not request.user.is_superuser: return redirect('home')
    thanh_vien = get_object_or_404(User, pk=pk)
    if request.method == 'POST':
        form = SuaThanhVienForm(request.POST, instance=thanh_vien)
        if form.is_valid(): form.save(); return redirect('quan_ly_thanh_vien')
    else: form = SuaThanhVienForm(instance=thanh_vien)
    return render(request, 'quanlysan/sua_thanh_vien.html', {'form': form, 'thanh_vien': thanh_vien})

@login_required
def cap_nhat_ho_so(request):
    if request.method == 'POST':
        form = CapNhatHoSoForm(request.POST, instance=request.user)
        if form.is_valid(): form.save(); return redirect('ho_so')
    else: form = CapNhatHoSoForm(instance=request.user)
    return render(request, 'quanlysan/form_chung.html', {'form': form, 'title': 'Cập nhật thông tin cá nhân'})

def gioi_thieu(request):
    info, created = ThongTinGioiThieu.objects.get_or_create(pk=1)
    return render(request, 'quanlysan/gioi_thieu.html', {'info': info})

@login_required
def sua_gioi_thieu(request):
    if not request.user.is_staff: raise PermissionDenied
    info, created = ThongTinGioiThieu.objects.get_or_create(pk=1)
    if request.method == 'POST':
        form = GioiThieuForm(request.POST, request.FILES, instance=info)
        if form.is_valid(): form.save(); return redirect('gioi_thieu')
    else: form = GioiThieuForm(instance=info)
    return render(request, 'quanlysan/form_chung.html', {'form': form, 'title': 'Sửa trang Giới thiệu'})

from rest_framework import viewsets
from rest_framework.permissions import IsAuthenticated
from .serializers import UserSerializer, DiaDiemSerializer, SanBongSerializer

class UserViewSet(viewsets.ModelViewSet):
    queryset = User.objects.all(); serializer_class = UserSerializer; permission_classes = [IsAuthenticated] 
class DiaDiemViewSet(viewsets.ModelViewSet):
    queryset = DiaDiem.objects.all(); serializer_class = DiaDiemSerializer; permission_classes = [IsAuthenticated] 
class SanBongViewSet(viewsets.ModelViewSet):
    queryset = SanBong.objects.all(); serializer_class = SanBongSerializer; permission_classes = [IsAuthenticated]

def error_404_view(request, exception): return render(request, 'quanlysan/404.html', status=404)
def error_403_view(request, exception=None): return render(request, 'quanlysan/403.html', status=403)

